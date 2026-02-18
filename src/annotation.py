import pandas as pd
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# --------------------------- FUNCTION DEFINITIONS -----------------------------

def build_annotated_regions(
    heatmap_matrix: pd.DataFrame,
    df_top_regions: pd.DataFrame,
    df_annot: pd.DataFrame,
    probe_col: str = "probe_ID",
    annot_probe_col: str = "Probe_ID",
) -> pd.DataFrame:
    """
    Join the probes that survived into heatmap_matrix with their differential
    stats (from df_top_regions) and genomic metadata (from df_annot).

    The join uses BOTH probe_ID AND tissue so that we only pull the row that
    was actually responsible for a probe being selected — not other tissues
    that probe may have ranked for.

    Parameters
    ----------
    heatmap_matrix : pd.DataFrame
        Tissues × regions matrix returned by create_heatmap_matrix().
        Columns are probe IDs; index is tissue names.
    df_top_regions : pd.DataFrame
        Long-format table with columns: probe_ID, tissue, target_meth,
        background_meth, diff.  Returned by extract_top_regions().
    df_annot : pd.DataFrame
        Annotation file with at minimum a Probe_ID column and genomic
        metadata columns (CpG_chrm, CpG_beg, CpG_end, gene name, etc.).
    probe_col : str
        Column name in df_top_regions that holds probe IDs.
    annot_probe_col : str
        Column name in df_annot that holds probe IDs.

    Returns
    -------
    pd.DataFrame
        Long-format annotated table sorted by tissue then diff (descending).
        Columns: tissue, probe_ID, target_meth, background_meth, diff,
                 + all annotation columns.
    """
    # probes present in heatmap_matrix (order preserved: left-to-right = region order)
    selected_probes    = heatmap_matrix.columns.tolist()
    selected_probe_set = set(selected_probes)

    # For each probe in heatmap_matrix we need to know WHICH tissue it was
    # selected for.  create_heatmap_matrix preserves the order produced by
    # extract_top_regions / groupby("tissue").head(max_per_tissue), so we can
    # reconstruct the tissue assignment by taking, for every probe that is in
    # selected_probe_set, the FIRST occurrence in df_top_regions (which
    # corresponds to the tissue that ranked it highest after uniqueness filtering).
    probe_to_tissue = (
        df_top_regions[df_top_regions[probe_col].isin(selected_probe_set)]
        .drop_duplicates(subset=probe_col, keep="first")
        [[probe_col, "tissue"]]
    )

    # pull the full stats row for each (probe, tissue) pair
    df_stats = df_top_regions.merge(
        probe_to_tissue,
        on=[probe_col, "tissue"],
        how="inner",
    )

    # merge with annotation
    df_annotated = df_stats.merge(
        df_annot,
        left_on=probe_col,
        right_on=annot_probe_col,
        how="left",
    )

    # drop duplicate probe column if annot uses a different name
    if annot_probe_col != probe_col and annot_probe_col in df_annotated.columns:
        df_annotated = df_annotated.drop(columns=[annot_probe_col])

    # sort: tissue alphabetically, then best-scoring probes first
    df_annotated = df_annotated.sort_values(
        ["tissue", "diff"], ascending=[True, False]
    ).reset_index(drop=True)

    print(f"Annotated regions shape: {df_annotated.shape}")
    print(f"Tissues represented: {df_annotated['tissue'].nunique()}")
    print(f"Unique probes: {df_annotated[probe_col].nunique()}")

    return df_annotated


def export_annotated_regions_excel(
    df_annotated: pd.DataFrame,
    output_path: Path,
    probe_col: str = "probe_ID",
) -> None:
    """
    Write df_annotated to a formatted Excel workbook (single sheet).

    Formatting:
    - Frozen header row with bold column names
    - Tissue column uses alternating light-blue banding per tissue block
    - diff, target_meth, background_meth formatted to 4 decimal places
    - Auto-sized columns (capped at 40)
    - A separate summary sheet with region counts per tissue
    """
    wb = openpyxl.Workbook()

    # ---- Sheet 1: All regions -----------------------------------------------
    ws = wb.active
    ws.title = "Annotated Regions"

    cols = df_annotated.columns.tolist()

    # Header
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(bottom=thin)

    for ci, col in enumerate(cols, start=1):
        cell            = ws.cell(row=1, column=ci, value=col)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = center
        cell.border     = border

    ws.freeze_panes           = "A2"
    ws.row_dimensions[1].height = 22

    # alternate band colours per tissue block
    band_colours = ["DCE6F1", "EBF3FB"]   # light blue alternating shades
    normal_font  = Font(name="Arial", size=10)
    number_fmt   = "0.0000"

    metric_cols = {"target_meth", "background_meth", "diff"}

    tissue_order = df_annotated["tissue"].unique().tolist()
    tissue_band  = {t: band_colours[i % 2] for i, t in enumerate(tissue_order)}

    for ri, row in enumerate(df_annotated.itertuples(index=False), start=2):
        row_dict   = row._asdict()
        tissue_val = row_dict["tissue"]
        fill       = PatternFill("solid", fgColor=tissue_band[tissue_val])

        for ci, col in enumerate(cols, start=1):
            val           = row_dict[col]
            cell          = ws.cell(row=ri, column=ci, value=val)
            cell.font     = normal_font
            cell.fill     = fill
            if col in metric_cols:
                cell.number_format = number_fmt

    # auto-size columns
    for ci, col in enumerate(cols, start=1):
        max_len = max(
            len(str(col)),
            df_annotated[col].astype(str).str.len().max(),
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 40)

    # ---- Sheet 2: Summary -------------------------------------------------
    ws_sum = wb.create_sheet("Region Counts")
    ws_sum.append(["Tissue", "N Regions", "Mean Diff", "Mean Target Meth", "Mean Background Meth"])

    sum_header_fill = PatternFill("solid", fgColor="375623")
    for cell in ws_sum[1]:
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = sum_header_fill
        cell.alignment = center

    summary = (
        df_annotated
        .groupby("tissue", sort=True)
        .agg(
            n_regions=(probe_col, "count"),
            mean_diff=("diff", "mean"),
            mean_target=("target_meth", "mean"),
            mean_bg=("background_meth", "mean"),
        )
        .reset_index()
    )

    for row in summary.itertuples(index=False):
        ws_sum.append([row.tissue, row.n_regions, row.mean_diff, row.mean_target, row.mean_bg])

    for ci in range(1, 6):
        ws_sum.column_dimensions[get_column_letter(ci)].width = 22
        if ci > 1:
            for ri in range(2, ws_sum.max_row + 1):
                ws_sum.cell(row=ri, column=ci).number_format = "0.0000" if ci > 2 else "0"

    wb.save(output_path)
    print(f"Excel saved → {output_path}")

# ------------------------------- END OF FUNCTION DEFINITIONS ----------------------------------
