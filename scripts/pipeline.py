# Import the modified CSV file for better visualization.
# Note that the added sufixes (".1", ".2", ..., ".25") is added by pandas when read by doing df = pd.read_csv(csv_file).
# The actual CSV does not contain those sufixes.
import pandas as pd
import numpy as np
from pathlib import Path
import matplotlib.pyplot as plt
import seaborn as sns
import sys
import warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from IPython.display import display, HTML
from scipy.cluster.hierarchy import linkage, dendrogram
from scipy.spatial.distance import pdist, squareform
from matplotlib.colors import LinearSegmentedColormap


DATA_DIR = Path.home() / "projects" / "data"

OUTPUT_DIR = Path.home() / "projects" / "mouse-DNA-project" / "results"
#OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

csv_file = DATA_DIR / "processed" / "GSE290585_SeSaMeBeta_MM285_BS_simplified_columns.csv" # path to mouse csv file

print(sys.executable)
print(csv_file.exists())

df = pd.read_csv(csv_file)

# -------------------------------- Filter df on NaN in the annotation csv file (originally from excel MM285) ----------------------------------------------
#Import Excel annotation file with metadata, identical probe_IDs as dataset -> we can filter on this and also map extracted regions back to the excel file.
annot_file = DATA_DIR / "original" / "12_02_2026_MM285.csv"
df_annot = pd.read_csv(annot_file, low_memory=False)

# drop any nan values from df_annot. We do not want to extract regions that contain no genomic information in the annotation file.
df_annot = df_annot.dropna(subset=["CpG_chrm"]).copy()

# make sure probe_id columns are strings
df["probe_ID"] = df["probe_ID"].astype(str)
df_annot["Probe_ID"] = df_annot["Probe_ID"].astype(str)

# create set of valid ids, for efficiency
valid_ids = set(df_annot["Probe_ID"])

# create mask based on valid_ids, and filter df using this mask + make it as independent dataframe by using copy (instead of accidentally creating a .view)
df = df[df["probe_ID"].isin(valid_ids)].copy() # Removing 7,415 rows, 296,070 rows -> 288,655 rows

print("Shape of df after NaN row filtering", df.shape)
# ----------------------------------------------------------------------------------------------------------------------------------------------------------

# --- Colormap used in plots ---
blue_yellow = LinearSegmentedColormap.from_list(
    "blue_yellow",
    ["#0000FF", "#ffff00"]
    )


# -------------------------------- df_cell_type ------------------------------
# Collapse cell type columns into one averaged column per cell type, resulting in 29 averaged cell types/columns
probe_ids = df["probe_ID"]
df_drop_ID = df.drop(columns="probe_ID")

# remove suffix .1 .2 etc. added by pandas
cell_type_labels = df_drop_ID.columns.str.split(".", n=1).str[0]

# collapse replicates
df_celltype = df_drop_ID.groupby(cell_type_labels, axis=1).mean()

# Reattach probe_ID
df_celltype.insert(0, "probe_ID", probe_ids)


# with pd.option_context("display.max_columns", None,
#                        "display.width", None):
#         display(df_celltype.head(20))

#df_celltype.shape
# ------------------------------------------------------------------------------


# --------------------------------- df_merged ----------------------------------
"""
After analysis of (725 region, 29 tissue) correlation matrix, we will do the following:

MERGE:
- Retina & Eye
- Brain_Cortex & Subcortinal_Brain
- Thymus & Blood & Spleen
- Tail & Ear & Skin

REMOVE:
- Optic_Nerve
- Sciatic_Nerve
(These are removed because cancer rarely metastasizes here)

This is done by filtering celltype_df and then merging celltypes. Expected: (525 regions, 21 tissues). HOWEVER, we will now use more than top25 regions -> Maybe top 100, 250 or even 500.
"""

# Drop columns that we don't need (or too low signal)
df_filtered_tissues = df_celltype.drop(columns=['Sciatic_Nerve', 'Optic_Nerve', 'Mammary_Glands'])

# Define "merge groups", the KEYS are the new merged tissue names.
merge_groups = {
    "Eye_Retina": ["Eye", "Retina"],
    "Brain_Cortex_Subcortical": ["Brain_Cortex", "Subcortical_Brain"],
    "Blood_Spleen_Thymus": ["Blood", "Spleen", "Thymus"],
    "Skin_Ears_Tail": ["Skin", "Ears", "Tail"],
}

# Build new merged df
df_merged = pd.DataFrame()
df_merged["probe_ID"] = df_filtered_tissues["probe_ID"] # copy over IDs

# add merged columns
for new_name, cols in merge_groups.items():
    df_merged[new_name] = df_filtered_tissues[cols].mean(axis=1)

# keep unmerged tissue as-is, flatten the merge_groups list
merged_cols_flat = [c for cols in merge_groups.values() for c in cols]

# keep non-merged columns
remaining_cols = []
for c in df_filtered_tissues.columns:
    if c not in merged_cols_flat and c != "probe_ID":
        remaining_cols.append(c)

# copy remaining columns into the merged df
df_merged[remaining_cols] = df_filtered_tissues[remaining_cols]

# print(df_merged.shape)
# print(df_merged.columns)

# with pd.option_context("display.max_columns", None,
#                        "display.width", None):
#         display(df_merged.head(20))
# ------------------------------------------------------------------------------


# --------------------------- FUNCTION DEFINITIONS -----------------------------
def build_tissue_diff_table(
    df: pd.DataFrame,
    target_threshold: float = 0.7,
    background_threshold: float = 0.8,
    diff_threshold: float = 0.3,
    use_filtering: bool = False,
    id_col: str = "probe_ID",
) -> pd.DataFrame:
    """
    Build a long-format table of tissue-specific differential methylation scores.

    This function iterates over each tissue (column) in a tissue-averaged
    methylation matrix and, for every genomic region (row), computes:
      - target methylation (in the current tissue)
      - background methylation (mean across all other tissues)
      - a differential score (background − target)

    For each tissue, regions are ranked by the differential score, optionally
    applying simple biological filtering thresholds. Results from all tissues
    are stacked into a single DataFrame suitable for marker selection,
    downstream filtering, or heatmap construction.

    Parameters
    ----------
    df_celltype : pandas.DataFrame
        DataFrame with one row per genomic region and one column per tissue,
        plus an identifier column (default: "probe_ID").
    target_threshold : float, optional
        Threshold for low methylation in the target tissue (used only if
        filtering conditions are enabled).
    background_threshold : float, optional
        Threshold for high methylation in non-target tissues (used only if
        filtering conditions are enabled).
    diff_threshold : float, optional
        Minimum differential score (background − target), currently optional
        and commented out by default.
    use_filtering : bool, optional
        If False, all regions are ranked purely by differential score.
        If True, biological filtering is applied before ranking.
    id_col : str, optional
        Name of the region identifier column.

    Returns
    -------
    all_tissues_results : list of DataFrames.
    """
    # -------------------------------
    # config
    # -------------------------------

    TARGET_THRESHOLD = target_threshold
    BACKGROUND_THRESHOLD = background_threshold
    DIFF_THRESHOLD = diff_threshold

    # -------------------------------
    # setup
    # -------------------------------

    probe_ids = df[id_col]
    tissues = df.columns.drop(id_col)

    all_tissues_results = []  # we will stack results here

    # -------------------------------
    # loop over tissues
    # -------------------------------

    for tissue in tissues:

        # target methylation (per row/region)
        target = df[tissue]

        # background methylation (mean of all OTHER tissues, per row)
        background = (
            df
            .drop(columns=[id_col, tissue])
            .mean(axis=1)
        )

        # difference
        diff = background - target

        # assemble per-tissue df
        tissue_df = pd.DataFrame({
            id_col: probe_ids,
            "tissue": tissue,
            "target_meth": target,
            "background_meth": background,
            "diff": diff
        })

        # switch between pure sorting on diff (use_filtering == FALSE) and filtering + sorting on diff (use_filtering == TRUE)
        if use_filtering:
            tissue_df_sorted = (
                tissue_df[
                    # (tissue_df["target_meth"] < TARGET_THRESHOLD) &
                    (tissue_df["background_meth"] > BACKGROUND_THRESHOLD)
                    # (tissue_df["diff"] > DIFF_THRESHOLD)
                ]
                .sort_values(by="diff", ascending=False)
            )
        
        else:
            tissue_df_sorted = tissue_df.sort_values(
            by="diff",
            ascending=False
            )

        # store results
        all_tissues_results.append(tissue_df_sorted)

    return all_tissues_results


def combine_tissue_results(all_tissues_results, verbose=False):
    """
    Concatenate per-tissue result tables into a single DataFrame.
    """

    final_df = pd.concat(all_tissues_results, ignore_index=True)

    if verbose:
        print("Final shape:", final_df.shape)
        display(final_df.head())

    return final_df


def extract_top_regions(
    all_tissues_results: list,
    top_n: int = 25
) -> pd.DataFrame:
    """
    Extract the top N ranked regions per tissue from per-tissue result tables.

    Parameters
    ----------
    all_tissues_results : list of pandas.DataFrame
        List of per-tissue DataFrames, each sorted by differential score.
        Typically returned by build_tissue_diff_table().
    top_n : int
        Number of top regions to extract per tissue.

    Returns
    -------
    pandas.DataFrame
        Long-format DataFrame containing the top N regions per tissue.
    """

    top_regions_per_tissue = []

    for tissue_df in all_tissues_results:
        top_n_df = tissue_df.head(top_n)
        top_regions_per_tissue.append(top_n_df)

    df_top_regions = pd.concat(top_regions_per_tissue, ignore_index=True)

    # Sanity check
    #print("Top regions shape:", df_top_regions.shape)

    return df_top_regions


def stats_top_regions(df_top_regions):
    """
    Function that extracts some summary statistics for df_top_regions.
    Requires df_top_regions structure with columns = ["target_meth", "background_meth", "diff"].
    """
    metrics_cols = ["target_meth", "background_meth", "diff"]
    df_metrics = df_top_regions[metrics_cols]

    summary_stats = df_metrics.describe().T
    summary_stats["range"] = summary_stats["max"] - summary_stats["min"]

    summary_stats = summary_stats.drop(columns=["count"]).T
    return summary_stats


def create_heatmap_matrix(df, df_top_regions, probe_col="probe_ID", region_mode="unique", max_per_tissue=None, verbose=False):
    """
    Create a heatmap-ready matrix (tissues x regions).
    Parameters:
    -----------
    region_mode : str
        - "all": use all regions (including duplicates)
        - "unique": keep only regions selected exactly once
    max_per_tissue : int, optional
        Maximum number of regions to keep per tissue. Only works with region_mode="unique".
        Regions are kept in their original ranking order from df_top_regions.
    verbose : bool
        If True, print detailed information about region selection
    """
    if region_mode not in {"all", "unique"}:
        raise ValueError("region_mode must be 'all' or 'unique'")
    
    if max_per_tissue is not None and region_mode != "unique":
        raise ValueError("max_per_tissue can only be used with region_mode='unique'. To resolve, set max_per_tissue=None.\n"
                        "Note: Using region_mode='all' will always return the same amount of regions as defined in top_N since there is no uniqueness constraint filtering.")
    
    if region_mode == "all":
        region_order = df_top_regions[probe_col].tolist()
    else:  # unique
        probe_counts = df_top_regions[probe_col].value_counts()
        unique_probes = probe_counts[probe_counts == 1].index
        region_order = (
            df_top_regions[df_top_regions[probe_col].isin(unique_probes)]
            [probe_col]
            .tolist()
        )
    
    # Calculate counts before capping
    counts_before = (
        df_top_regions[df_top_regions[probe_col].isin(region_order)]
        ["tissue"]
        .value_counts()
        .rename("n_regions")
        .to_frame()
        .sort_index()
    )
    
    # Apply max_per_tissue cap if specified
    if max_per_tissue is not None:
        # Check for tissues below max_per_tissue threshold
        tissues_below_max = counts_before[counts_before["n_regions"] < max_per_tissue]
        if not tissues_below_max.empty:
            warning_msg = "Some tissues have fewer regions than max_per_tissue:\n"
            for tissue, row in tissues_below_max.iterrows():
                warning_msg += f"  - {tissue}: {row['n_regions']} regions\n"
            warnings.warn(warning_msg, UserWarning)
        
        # Filter to unique regions and apply cap while preserving order
        df_filtered = df_top_regions[df_top_regions[probe_col].isin(region_order)]
        
        # Group by tissue and take first max_per_tissue regions (preserves df_top_regions order)
        capped_regions = (
            df_filtered
            .groupby("tissue", sort=False)
            .head(max_per_tissue)
            [probe_col]
            .tolist()
        )
        region_order = capped_regions
    
    # Calculate final counts
    counts = (
        df_top_regions[df_top_regions[probe_col].isin(region_order)]
        ["tissue"]
        .value_counts()
        .rename("n_regions")
        .to_frame()
        .sort_index()
    )
    
    # Reorder counts to match heatmap x-axis order
    ordered_tissues = (
        df_top_regions
        .loc[df_top_regions[probe_col].isin(region_order), "tissue"]
        .drop_duplicates()
        .tolist()
    )
    counts = counts.loc[ordered_tissues]
    
    if verbose:
        print(f"Region mode: {region_mode}")
        
        if max_per_tissue is not None:
            print(f"\n--- BEFORE capping (max_per_tissue={max_per_tissue}) ---")
            print(f"Total regions: {len(counts_before)}")
            display(counts_before)
            
            print(f"\n--- AFTER capping ---")
            print(f"Total regions: {len(region_order)}")
            display(counts)
        else:
            print(f"Total regions used: {len(region_order)}")
            display(counts)
    
    heatmap_df = (
        df
        .set_index(probe_col)
        .loc[region_order]
    )
    
    return heatmap_df.T, counts


def plot_heatmap(heatmap_matrix, counts, top_n=None, title=None, figsize=(20, 8)):
    """
    Plot methylation heatmap with automatic tissue block labels.

    Parameters
    ----------
    heatmap_matrix : pd.DataFrame
        DataFrame of shape (tissues × regions)
    counts : pd.DataFrame
        DataFrame with index = tissue names (in plot order)
        and column 'n_regions' giving number of regions per tissue
    """

    # extract labels and block sizes from counts
    # counts is a df with index (cell type) and n_regions in each tissue
    tissue_labels = counts.index.tolist()
    block_sizes = counts["n_regions"].values

    # compute xtick positions (center of each tissue block)
    xtick_positions = []
    current_pos = 0

    for size in block_sizes:
        xtick_positions.append(current_pos + size / 2)
        current_pos += size

    # plot
    plt.figure(figsize=figsize)

    ax = sns.heatmap(
        heatmap_matrix,
        cmap=blue_yellow,
        vmin=0,
        vmax=1,
        xticklabels=False,
        yticklabels=True,
        cbar_kws={"shrink": 0.75, "pad": 0.03}
    )

    ax.set_xticks(xtick_positions)
    ax.set_xticklabels(tissue_labels, rotation=90)
    ax.set_ylabel("Tissues")

    if top_n is not None:
        ax.set_xlabel(f"Tissue-specific regions (top {top_n} per tissue)")
    else:
        ax.set_xlabel("Tissue-specific regions")

    if title is not None:
        ax.set_title(title)

    plt.show()


def compute_tissue_correlation(heatmap_matrix, method="pearson"):
    """
    Compute tissue–tissue correlation matrix. Shared computation function, avoids recomputing T.corr
    """
    return heatmap_matrix.T.corr(method=method)


def plot_tissue_correlation(tissue_corr, figsize=(10, 8), title=None):
    """
    Plot tissue correlation, colormap between 0 and 1.
    """
    plt.figure(figsize=figsize)
    sns.heatmap(
        tissue_corr,
        cmap="coolwarm",
        vmin=0,
        vmax=1,
        square=True,
        linewidths=0.5,
        cbar_kws={"shrink": 0.75, "pad": 0.03}
    )
    plt.title(title or "Tissue–tissue methylation correlation")
    plt.tight_layout()
    plt.show()


def plot_tissue_dendrogram(tissue_corr, method="average", figsize=(10, 6), title=None):
    """
    Plot Dendrogram
    """
    # hierarchical clustering is based on distances, here we convert tissue_corr to distances giving us:
    # correlation = 1  -> distance = 0 (identical) 
    # correlation = 0  -> distance = 1 (unrelated)
    # correlation = -1 -> distance = 2 (opposite)
    distance_matrix = 1 - tissue_corr
    condensed_dist = squareform(distance_matrix.values)

    Z = linkage(condensed_dist, method=method)
    plt.figure(figsize=figsize)
    dendrogram(
        Z,
        labels=tissue_corr.index.tolist(),
        leaf_rotation=90
    )

    plt.ylabel("Distance (1 - correlation)")
    plt.title(title or "Hierarchical clustering of tissues")
    plt.tight_layout()
    plt.show()
    return Z


def plot_clustered_tissue_correlation(tissue_corr, figsize=(10, 10), title=None):
    """
    Plot clustered tissue correlation.
    """
    sns.clustermap(
        tissue_corr,
        cmap="coolwarm",
        vmin=0,
        vmax=1,
        linewidths=0.5,
        figsize=figsize,
        cbar_kws={"shrink": 0.75, "pad": 0.03}
    )
    if title:
        plt.suptitle(title, y=1.02)


def build_annotated_regions(
    heatmap_matrix: pd.DataFrame,
    df_top_regions: pd.DataFrame,
    df_annot: pd.DataFrame,
    probe_col: str = "probe_ID",
    annot_probe_col: str = "Probe_ID",
) -> pd.DataFrame:
    """
    Join the probes that survived into heatmap_matrix with their differential
    stats (from df_top_regions) and genomic metadata (from df_annot).

    The join uses BOTH probe_ID AND tissue so that we only pull the row that
    was actually responsible for a probe being selected — not other tissues
    that probe may have ranked for.

    Parameters
    ----------
    heatmap_matrix : pd.DataFrame
        Tissues × regions matrix returned by create_heatmap_matrix().
        Columns are probe IDs; index is tissue names.
    df_top_regions : pd.DataFrame
        Long-format table with columns: probe_ID, tissue, target_meth,
        background_meth, diff.  Returned by extract_top_regions().
    df_annot : pd.DataFrame
        Annotation file with at minimum a Probe_ID column and genomic
        metadata columns (CpG_chrm, CpG_beg, CpG_end, gene name, etc.).
    probe_col : str
        Column name in df_top_regions that holds probe IDs.
    annot_probe_col : str
        Column name in df_annot that holds probe IDs.

    Returns
    -------
    pd.DataFrame
        Long-format annotated table sorted by tissue then diff (descending).
        Columns: tissue, probe_ID, target_meth, background_meth, diff,
                 + all annotation columns.
    """
    # probes present in heatmap_matrix (order preserved: left-to-right = region order)
    selected_probes = heatmap_matrix.columns.tolist()
    selected_probe_set = set(selected_probes)

    # For each probe in heatmap_matrix we need to know WHICH tissue it was
    # selected for.  create_heatmap_matrix preserves the order produced by
    # extract_top_regions / groupby("tissue").head(max_per_tissue), so we can
    # reconstruct the tissue assignment by taking, for every probe that is in
    # selected_probe_set, the FIRST occurrence in df_top_regions (which
    # corresponds to the tissue that ranked it highest after uniqueness filtering).
    probe_to_tissue = (
        df_top_regions[df_top_regions[probe_col].isin(selected_probe_set)]
        .drop_duplicates(subset=probe_col, keep="first")
        [[probe_col, "tissue"]]
    )

    # pull the full stats row for each (probe, tissue) pair
    df_stats = df_top_regions.merge(
        probe_to_tissue,
        on=[probe_col, "tissue"],
        how="inner",
    )

    #merge with annotation
    df_annotated = df_stats.merge(
        df_annot,
        left_on=probe_col,
        right_on=annot_probe_col,
        how="left",
    )

    # drop duplicate probe column if annot uses a different name
    if annot_probe_col != probe_col and annot_probe_col in df_annotated.columns:
        df_annotated = df_annotated.drop(columns=[annot_probe_col])

    # sort: tissue alphabetically, then best-scoring probes first
    df_annotated = df_annotated.sort_values(
        ["tissue", "diff"], ascending=[True, False]
    ).reset_index(drop=True)

    print(f"Annotated regions shape: {df_annotated.shape}")
    print(f"Tissues represented: {df_annotated['tissue'].nunique()}")
    print(f"Unique probes: {df_annotated[probe_col].nunique()}")

    return df_annotated


def export_annotated_regions_excel(
    df_annotated: pd.DataFrame,
    output_path: Path,
    probe_col: str = "probe_ID",
) -> None:
    """
    Write df_annotated to a formatted Excel workbook (single sheet).

    Formatting:
    - Frozen header row with bold column names
    - Tissue column uses alternating light-blue banding per tissue block
    - diff, target_meth, background_meth formatted to 4 decimal places
    - Auto-sized columns (capped at 40)
    - A separate summary sheet with region counts per tissue
    """
    wb = openpyxl.Workbook()

    # ---- Sheet 1: All regions -----------------------------------------------
    ws = wb.active
    ws.title = "Annotated Regions"

    cols = df_annotated.columns.tolist()

    # Header
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(bottom=thin)

    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    # alternate band colours per tissue block
    band_colours = ["DCE6F1", "EBF3FB"]   # light blue alternating shades
    normal_font = Font(name="Arial", size=10)
    number_fmt = "0.0000"

    metric_cols = {"target_meth", "background_meth", "diff"}

    tissue_order = df_annotated["tissue"].unique().tolist()
    tissue_band = {t: band_colours[i % 2] for i, t in enumerate(tissue_order)}

    for ri, row in enumerate(df_annotated.itertuples(index=False), start=2):
        tissue_val = getattr(row, "tissue")
        fill = PatternFill("solid", fgColor=tissue_band[tissue_val])

        for ci, col in enumerate(cols, start=1):
            val = getattr(row, col.replace(" ", "_"))
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = normal_font
            cell.fill = fill
            if col in metric_cols:
                cell.number_format = number_fmt

    # auto-size columns
    for ci, col in enumerate(cols, start=1):
        max_len = max(
            len(str(col)),
            df_annotated[col].astype(str).str.len().max(),
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 40)

    # ---- Sheet 2: Summary -------------------------------------------------
    ws_sum = wb.create_sheet("Region Counts")
    ws_sum.append(["Tissue", "N Regions", "Mean Diff", "Mean Target Meth", "Mean Background Meth"])

    sum_header_fill = PatternFill("solid", fgColor="375623")
    for cell in ws_sum[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = sum_header_fill
        cell.alignment = center

    summary = (
        df_annotated
        .groupby("tissue", sort=True)
        .agg(
            n_regions=(probe_col, "count"),
            mean_diff=("diff", "mean"),
            mean_target=("target_meth", "mean"),
            mean_bg=("background_meth", "mean"),
        )
        .reset_index()
    )

    for row in summary.itertuples(index=False):
        ws_sum.append([row.tissue, row.n_regions, row.mean_diff, row.mean_target, row.mean_bg])

    for ci in range(1, 6):
        ws_sum.column_dimensions[get_column_letter(ci)].width = 22
        if ci > 1:
            for ri in range(2, ws_sum.max_row + 1):
                ws_sum.cell(row=ri, column=ci).number_format = "0.0000" if ci > 2 else "0"

    wb.save(output_path)
    print(f"Excel saved → {output_path}")
# ------------------------------- END OF FUNCTION DEFINITIONS ----------------------------------


# ---------------------------------- Pipeline Configuration ------------------------------------
# input dataframe
DATAFRAME = df_merged            # tissue-averaged beta matrix (probe_ID + tissues)

# use filtering + diff ranking (True) VS. Only diff ranking (False)
USE_FILTERING = True               # passed to build_tissue_diff_table
USE_VERBOSE = True                 # display the distribution of regions in cell types.
MAX_PER_TISSUE = 50

# region selection
TOP_N = 200                        # number of top regions per tissue
REGION_MODE = "unique"                # "all" or "unique"

# correlation / clustering
CORR_METHOD = "pearson"            # "pearson" or "spearman"
CLUSTER_METHOD = "average"         # linkage method

# plotting
HEATMAP_TITLE = "Mouse methylation atlas heatmap"
CORR_TITLE = "Tissue–tissue methylation correlation"
DENDRO_TITLE = "Hierarchical clustering of tissues (methylation correlation)"
CLUSTERED_CORR_TITLE = "Clustered tissue–tissue methylation correlation"
# ---------------------------------------------------------------------------------------------


# ----------------------------------------- PIPELINE ------------------------------------------
""" FULL PIPELINE - 20 TISSUES """
# Step 1: compute per-tissue differential tables
all_tissues_results = build_tissue_diff_table(
    df=DATAFRAME,
    use_filtering=USE_FILTERING
)

# Step 2: extract top N regions per tissue
df_top_regions = extract_top_regions(
    all_tissues_results,
    top_n=TOP_N
)

# Step 3: create heatmap matrix (tissues × regions)
heatmap_matrix, regions_per_tissue_count = create_heatmap_matrix(
    df=DATAFRAME,
    df_top_regions=df_top_regions,
    region_mode=REGION_MODE,
    max_per_tissue=MAX_PER_TISSUE,
    verbose=USE_VERBOSE
)

# Step 4: plot methylation heatmap
plot_heatmap(
    heatmap_matrix,
    regions_per_tissue_count,
    top_n=TOP_N,
    title=f"{HEATMAP_TITLE} (top {TOP_N} regions per tissue, truncated at {MAX_PER_TISSUE} regions per tissue) - {REGION_MODE} regions"
)

# Step 5: compute tissue–tissue correlation matrix
tissue_corr = compute_tissue_correlation(
    heatmap_matrix,
    method=CORR_METHOD
)

# Step 6: plot tissue correlation heatmap
plot_tissue_correlation(
    tissue_corr,
    title=CORR_TITLE
)

# Step 7: plot tissue dendrogram (hierarchical clustering)
Z = plot_tissue_dendrogram(
    tissue_corr,
    method=CLUSTER_METHOD,
    title=DENDRO_TITLE
)

# Step 8: plot clustered tissue correlation heatmap
plot_clustered_tissue_correlation(
    tissue_corr,
    title=CLUSTERED_CORR_TITLE
)

# Step 9: join surviving probes (after unique filtering) with metadata annotations in single df
df_annot_final = build_annotated_regions(
    heatmap_matrix=heatmap_matrix,
    df_top_regions=df_top_regions,
    df_annot= df_annot
)

# Step 10: export the joint probe and metadata df to an excel file, saved to output_path.
output_path = OUTPUT_DIR / "mouse_methylation_markers.xlsx"
#print(output_path)
#print(output_path.exists())
export_annotated_regions_excel(
    df_annotated=df_annot_final,
    output_path=output_path
)
# ---------------------------------------------------------------------------------------------


# -------------------------------- VALIDATION & VALIDATION PLOTS ------------------------------
# """
# PCA QC Check - Do my extracted regions in heatmap_matrix encode useful information?
# """

# from sklearn.decomposition import PCA
# from sklearn.preprocessing import StandardScaler

# # convert df to numpy array
# X = heatmap_matrix.values
# #extract tissues
# tissues = heatmap_matrix.index

# # try both with scaled and not scaled. We center in both cases.
# # X_centered = StandardScaler(with_std=False).fit_transform(X)
# X_scaled = StandardScaler(with_std=True).fit_transform(X)


# pca = PCA(n_components=10)
# #X_pca_centered = pca.fit_transform(X_centered)
# X_pca = pca.fit_transform(X_scaled)

# explained = pca.explained_variance_ratio_
# cumulative = np.cumsum(explained)

# print("Explained variance (first 10 PCs):")
# for i in range(10):
#     print(f"PC{i+1}: {explained[i]*100:.1f}%")

# print("\nCumulative variance (first 10 PCs):")
# for i in range(10):
#     print(f"PC1–PC{i+1}: {cumulative[i]*100:.1f}%")


# # Plot PC1 vs PC2
# plt.figure(figsize=(8, 6))
# plt.scatter(X_pca[:, 0], X_pca[:, 1])

# for i, t in enumerate(tissues):
#     plt.text(X_pca[i, 0], X_pca[i, 1], t, fontsize=6)

# plt.xlabel(f"PC1 ({pca.explained_variance_ratio_[0]*100:.1f}%)")
# plt.ylabel(f"PC2 ({pca.explained_variance_ratio_[1]*100:.1f}%)")
# plt.title("PCA on atlas-selected regions")
# plt.tight_layout()
# plt.show()



# # plot PC2 vs PC3
# plt.figure(figsize=(8, 6))
# plt.scatter(X_pca[:, 1], X_pca[:, 2])

# for i, t in enumerate(tissues):
#     plt.text(X_pca[i, 1], X_pca[i, 2], t, fontsize=9)

# plt.xlabel(f"PC2 ({pca.explained_variance_ratio_[1]*100:.1f}%)")
# plt.ylabel(f"PC3 ({pca.explained_variance_ratio_[2]*100:.1f}%)")
# plt.title("PCA on atlas-selected regions")
# plt.tight_layout()
# plt.show()


# # plot PC1 vs PC3
# plt.figure(figsize=(8, 6))
# plt.scatter(X_pca[:, 0], X_pca[:, 2])

# for i, t in enumerate(tissues):
#     plt.text(X_pca[i, 0], X_pca[i, 2], t, fontsize=9)

# plt.xlabel(f"PC1 ({pca.explained_variance_ratio_[0]*100:.1f}%)")
# plt.ylabel(f"PC3 ({pca.explained_variance_ratio_[2]*100:.1f}%)")
# plt.title("PCA on atlas-selected regions")
# plt.tight_layout()
# plt.show()


# # plot PC3 vs PC4
# plt.figure(figsize=(8, 6))
# plt.scatter(X_pca[:, 2], X_pca[:, 3])

# for i, t in enumerate(tissues):
#     plt.text(X_pca[i, 2], X_pca[i, 3], t, fontsize=9)

# plt.xlabel(f"PC3 ({pca.explained_variance_ratio_[2]*100:.1f}%)")
# plt.ylabel(f"PC4 ({pca.explained_variance_ratio_[3]*100:.1f}%)")
# plt.title("PCA on atlas-selected regions")
# plt.tight_layout()
# plt.show()


# """
# TEST: Remove testis to see if it was dominating PC1 - if it was "testis vs all"
# """
# # remove testis - only diagnostic
# heatmap_matrix_no_testis = heatmap_matrix.drop(index="Testis")

# print("Original shape:", heatmap_matrix.shape)
# print("No Testis shape:", heatmap_matrix_no_testis.shape)


# X_no_testis = heatmap_matrix_no_testis.values
# tissues_no_testis = heatmap_matrix_no_testis.index


# X_no_testis = StandardScaler().fit_transform(X_no_testis)

# pca_no_testis = PCA()
# X_pca_no_testis = pca_no_testis.fit_transform(X_no_testis)


# explained = pca_no_testis.explained_variance_ratio_
# cumulative = np.cumsum(explained)

# print("Explained variance (first 10 PCs):")
# for i in range(10):
#     print(f"PC{i+1}: {explained[i]*100:.1f}%")

# print("\nCumulative variance (first 10 PCs):")
# for i in range(10):
#     print(f"PC1–PC{i+1}: {cumulative[i]*100:.1f}%")


# # No testis PC1 vs PC2
# plt.figure(figsize=(7, 6))
# plt.scatter(X_pca_no_testis[:, 0], X_pca_no_testis[:, 1])

# for i, t in enumerate(tissues_no_testis):
#     plt.text(X_pca_no_testis[i, 0], X_pca_no_testis[i, 1], t, fontsize=9)

# plt.xlabel(f"PC1 ({explained[0]*100:.1f}%)")
# plt.ylabel(f"PC2 ({explained[1]*100:.1f}%)")
# plt.title("PCA (testis removed)")
# plt.tight_layout()
# plt.show()


# # No testis PC2 vs PC3
# plt.figure(figsize=(7, 6))
# plt.scatter(X_pca_no_testis[:, 1], X_pca_no_testis[:, 2])

# for i, t in enumerate(tissues_no_testis):
#     plt.text(X_pca_no_testis[i, 1], X_pca_no_testis[i, 2], t, fontsize=9)

# plt.xlabel(f"PC2 ({explained[1]*100:.1f}%)")
# plt.ylabel(f"PC3 ({explained[2]*100:.1f}%)")
# plt.title("PCA (testis removed)")
# plt.tight_layout()
# plt.show()

# Some more sanity checks:
list_probes = heatmap_matrix.columns.to_list()
list_probes_set = set(list_probes)

mask = df_top_regions[df_top_regions["probe_ID"].isin(list_probes_set)]
print(mask.nunique())